import openpyxl
import json
import re
from datetime import date, datetime
import argparse
from pathlib import Path

class MenuDataExtractor:
    """Excel verilerini JSON formatına dönüştüren Domain Service."""
    
    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = openpyxl.load_workbook(file_path, data_only=True)
        self.menu_registry = {}

    @staticmethod
    def _tr_upper_ascii(text: str) -> str:
        """Türkçe karakterleri ASCII'ye yaklaştırarak sheet name eşleştirmeyi sağlamlaştırır."""
        if text is None:
            return ""
        t = str(text).strip().upper()
        # Yaygın Türkçe harfleri sadeleştir.
        t = (
            t.replace("Ç", "C")
            .replace("Ğ", "G")
            .replace("İ", "I")
            .replace("İ", "I")
            .replace("Ö", "O")
            .replace("Ş", "S")
            .replace("Ü", "U")
        )
        # Bazı terminal/encoding durumlarında görülen bozuk karakterleri yok say.
        t = t.replace("�", "")
        return t

    @staticmethod
    def _parse_date_text(text: str) -> str | None:
        """Excel hücresindeki tarih yazısını ISO (YYYY-MM-DD) formatına çevirir."""
        if not text:
            return None
        t = str(text).strip()
        if not t:
            return None
        # Bazı dosyalarda "09.04..2026" gibi hatalı yazımlar var.
        t = re.sub(r"([./-])\1+", r"\1", t)

        m = re.search(r"\b(\d{4})-(\d{2})-(\d{2})\b", t)
        if m:
            try:
                d = date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            except ValueError:
                return None
            return d.isoformat()

        m = re.search(r"\b(\d{1,2})[./-]+(\d{1,2})[./-]+(\d{2,4})\b", t)
        if not m:
            return None

        day = int(m.group(1))
        month = int(m.group(2))
        year = int(m.group(3))
        if year < 100:
            year = 2000 + year

        try:
            d = date(year, month, day)
        except ValueError:
            return None
        return d.isoformat()

    def _find_sheet(self, kind: str):
        names = list(self.wb.sheetnames)
        if kind == "kahvalti":
            for name in names:
                if self._tr_upper_ascii(name) == "KAHVALTI":
                    return self.wb[name]
            for name in names:
                if "KAHVALTI" in self._tr_upper_ascii(name) or "KAH" in self._tr_upper_ascii(name):
                    return self.wb[name]
            raise KeyError("KAHVALTI sayfası bulunamadı")

        if kind == "aksam":
            for name in names:
                if self._tr_upper_ascii(name) in ("AKSAM MENU", "AKSAM MENUSU", "AKSAM MENÜ", "AKSAM MENU"):
                    return self.wb[name]
            for name in names:
                up = self._tr_upper_ascii(name)
                if "AKSAM" in up and ("MENU" in up or "MEN" in up):
                    return self.wb[name]
            raise KeyError("AKŞAM MENÜ sayfası bulunamadı")

        raise ValueError(f"Bilinmeyen sheet kind: {kind}")

    def _ensure_date_entry(self, date_str: str):
        if date_str not in self.menu_registry:
            self.menu_registry[date_str] = {"kahvalti": [], "aksam": []}

    def _format_calories(self, value):
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        if re.search(r"\bkcal\b", text, flags=re.IGNORECASE):
            # "120 kcal" gibi değerleri tek biçime indir.
            text = re.sub(r"\s*kcal\b", "", text, flags=re.IGNORECASE).strip()
        return f"{text} kcal"

    def _iter_date_anchors(self, sheet, max_rows: int, max_cols: int):
        """Sayfadaki datetime hücrelerini (tarih anchor) döndürür."""
        seen = set()
        for r in range(1, max_rows + 1):
            for c in range(1, max_cols + 1):
                cell = sheet.cell(row=r, column=c)
                v = cell.value

                date_str = None
                if isinstance(v, datetime):
                    date_str = v.date().isoformat()
                elif isinstance(v, date):
                    date_str = v.isoformat()
                elif isinstance(v, (int, float)) and getattr(cell, "is_date", False):
                    # Bazı Excel'ler tarihi sayısal seri olarak tutuyor.
                    try:
                        from openpyxl.utils.datetime import from_excel

                        dt = from_excel(v, self.wb.epoch)
                        date_str = dt.date().isoformat()
                    except Exception:
                        date_str = None
                elif isinstance(v, str):
                    date_str = self._parse_date_text(v)

                if not date_str:
                    continue

                key = (date_str, r, c)
                if key in seen:
                    continue
                seen.add(key)
                yield date_str, r, c

    def _get_formatted_date(self, cell_value):
        if isinstance(cell_value, datetime):
            return cell_value.strftime('%Y-%m-%d')
        if isinstance(cell_value, str):
            return self._parse_date_text(cell_value)
        return None

    def extract_aksam(self):
        sheet = self._find_sheet("aksam")

        def _norm_name(value):
            # CSV kaynaklarında genellikle CRLF kullanılıyor; Excel hücreleri LF döndürebiliyor.
            return str(value).strip().replace('\r\n', '\n').replace('\n', '\r\n')

        # Akşam: tarih hücresinden itibaren +2..+7 satırlarında (6 satır) iki blok var:
        # Ana Menü: (c, c+1) | Salatbar: (c+2, c+3)
        for date_str, r_start, c_start in self._iter_date_anchors(sheet, max_rows=250, max_cols=60):
            self._ensure_date_entry(date_str)
            items = []
            for i in range(2, 8):
                name = sheet.cell(row=r_start + i, column=c_start).value
                cal = sheet.cell(row=r_start + i, column=c_start + 1).value
                if name and str(name).strip().upper() != 'TOPLAM':
                    items.append({
                        "category": "Ana Menü",
                        "name": _norm_name(name),
                        "calories": self._format_calories(cal)
                    })

                s_name = sheet.cell(row=r_start + i, column=c_start + 2).value
                s_cal = sheet.cell(row=r_start + i, column=c_start + 3).value
                if s_name and str(s_name).strip().upper() != 'TOPLAM':
                    items.append({
                        "category": "Salatbar",
                        "name": _norm_name(s_name),
                        "calories": self._format_calories(s_cal)
                    })

            if items:
                self.menu_registry[date_str]["aksam"] = items

    def extract_kahvalti(self):
        sheet = self._find_sheet("kahvalti")

        # Kahvaltı: tarih hücresinden itibaren +1..+7 satırlarında (7 satır) isim+kalori blokları var.
        for date_str, r_start, c_start in self._iter_date_anchors(sheet, max_rows=250, max_cols=60):
            self._ensure_date_entry(date_str)
            items = []
            for i in range(1, 8):
                name = sheet.cell(row=r_start + i, column=c_start).value
                cal = sheet.cell(row=r_start + i, column=c_start + 1).value
                if name and str(name).strip().upper() != 'TOPLAM':
                    items.append({
                        "category": "Kahvaltılık",
                        "name": str(name).strip().replace('\r\n', '\n').replace('\n', '\r\n'),
                        "calories": self._format_calories(cal)
                    })
            if items:
                self.menu_registry[date_str]["kahvalti"] = items

    def save_json(self, output_path):
        # Tarihe göre sıralayıp kaydedelim
        sorted_data = dict(sorted(self.menu_registry.items()))
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(sorted_data, f, ensure_ascii=False, indent=2)

    def compare_with_menu_org(self, menu_org_path: str = 'menu-org.json'):
        """menu_registry ile menu-org.json arasındaki farkları raporlar."""
        path = Path(menu_org_path)
        if not path.exists():
            return {
                "menu_org_exists": False,
                "missing_in_org": sorted(self.menu_registry.keys()),
                "missing_in_registry": [],
                "diff_kahvalti": [],
                "diff_aksam": [],
            }

        org = json.loads(path.read_text(encoding='utf-8'))

        def _norm(obj):
            return json.dumps(obj, ensure_ascii=False, sort_keys=True)

        missing_in_org = []
        missing_in_registry = []
        diff_kahvalti = []
        diff_aksam = []

        for date_str, day in self.menu_registry.items():
            if date_str not in org:
                missing_in_org.append(date_str)
                continue
            if _norm(day.get('kahvalti', [])) != _norm(org[date_str].get('kahvalti', [])):
                diff_kahvalti.append(date_str)
            if _norm(day.get('aksam', [])) != _norm(org[date_str].get('aksam', [])):
                diff_aksam.append(date_str)

        for date_str in org.keys():
            if date_str not in self.menu_registry:
                missing_in_registry.append(date_str)

        return {
            "menu_org_exists": True,
            "missing_in_org": sorted(missing_in_org),
            "missing_in_registry": sorted(missing_in_registry),
            "diff_kahvalti": sorted(diff_kahvalti),
            "diff_aksam": sorted(diff_aksam),
        }

    def sync_menu_org(
        self,
        menu_org_path: str = 'menu-org.json',
        only_if_non_empty: bool = True,
        add_missing_dates: bool = False,
    ):
        """menu-org.json içinde kahvalti/aksam alanlarını menu_registry'ye göre günceller.

        Varsayılan davranış güvenli olacak şekilde sadece mevcut günleri günceller.
        add_missing_dates=True verilirse Excel'de olup menu-org.json'da olmayan günleri de ekler.
        """
        path = Path(menu_org_path)
        org = json.loads(path.read_text(encoding='utf-8')) if path.exists() else {}

        for date_str, day in self.menu_registry.items():
            if date_str not in org:
                if not add_missing_dates:
                    continue
                org[date_str] = {"kahvalti": [], "aksam": []}

            for key in ("kahvalti", "aksam"):
                new_items = day.get(key, [])
                if only_if_non_empty and not new_items:
                    continue
                org[date_str][key] = new_items

        sorted_data = dict(sorted(org.items()))
        path.write_text(json.dumps(sorted_data, ensure_ascii=False, indent=2), encoding='utf-8')

# --- Uygulama Noktası ---
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Excel menülerini JSON'a çevirir ve kontrol raporu üretir.")
    parser.add_argument("--excel", default="nisan.xlsx", help="Kaynak Excel dosyası")
    parser.add_argument("--out", default="menu.json", help="Üretilecek menu.json yolu")
    parser.add_argument("--menu-org", default="menu-org.json", help="Karşılaştırılacak menu-org.json yolu")
    parser.add_argument("--sync-menu-org", action="store_true", help="menu-org.json'u (non-empty) güncelle")
    parser.add_argument("--add-missing-dates", action="store_true", help="Sync sırasında menu-org.json'da olmayan günleri de ekle")
    args = parser.parse_args()

    extractor = MenuDataExtractor(args.excel)
    extractor.extract_aksam()
    extractor.extract_kahvalti()
    extractor.save_json(args.out)

    report = extractor.compare_with_menu_org(args.menu_org)
    print(f"menu.json üretildi: {args.out}")
    print(f"Toplam gün: {len(extractor.menu_registry)}")
    print(f"Fark (kahvalti): {len(report['diff_kahvalti'])} | Fark (aksam): {len(report['diff_aksam'])}")

    def _preview(values, limit: int = 12):
        if len(values) <= limit:
            return values
        return values[:limit] + [f"... (+{len(values) - limit})"]

    if report["diff_kahvalti"]:
        print(f"Kahvaltı farklı günler: {_preview(report['diff_kahvalti'])}")
    if report["diff_aksam"]:
        print(f"Akşam farklı günler: {_preview(report['diff_aksam'])}")
    if report["missing_in_org"]:
        print(f"menu-org.json içinde eksik gün sayısı: {len(report['missing_in_org'])} | örnek: {_preview(report['missing_in_org'])}")
    if report["missing_in_registry"]:
        print(f"Excel'de olmayan ama menu-org.json'da olan gün sayısı: {len(report['missing_in_registry'])} | örnek: {_preview(report['missing_in_registry'])}")

    if args.sync_menu_org:
        extractor.sync_menu_org(
            args.menu_org,
            only_if_non_empty=True,
            add_missing_dates=args.add_missing_dates,
        )
        if args.add_missing_dates:
            print("menu-org.json güncellendi + eksik günler eklendi (boş listeler yazılmadı).")
        else:
            print("menu-org.json güncellendi (sadece mevcut günler; boş listeler yazılmadı).")

        post = extractor.compare_with_menu_org(args.menu_org)
        print(f"Sync sonrası fark (kahvalti): {len(post['diff_kahvalti'])} | fark (aksam): {len(post['diff_aksam'])}")